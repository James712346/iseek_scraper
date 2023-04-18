from openpyxl import load_workbook
from asyncio import run
from tortoise import Tortoise, exceptions 
from models import Graphs



async def start(DatabaseUrl):

    await Tortoise.init(db_url = DatabaseUrl,modules={"models": ["models"]} )

    workbook = load_workbook(filename="VNNI CVC.xlsx")

    for nsheet in workbook.sheetnames[1:]:
        sheet = workbook[nsheet]
        RowsToUpdate = []
        for row in sheet.iter_rows(min_row=2, max_col=17, values_only=True):
            Poicode = row[5].split("-")[0].rstrip()
            graph = None
            if len(Poicode) == 4:
                graph = await Graphs.get_or_none(POI_code=Poicode)
            if not graph:
                graph = await Graphs.get_or_none(NBN_CVC=row[8])
            if graph:
                graph.AGVL = row[1]
                graph.Optus_STag = row[2]
                graph.NBN_STag = row[3]
                graph.VTag_VLAN_ID = row[4]
                graph.CSA =row[6]
                graph.OPTUS_CVC = row[7]
                graph.NBN_CVC = row[8]
                graph.OON = row[9]
                graph.ORD_VLINK = row[10]
                graph.ORD_CVC = row [11]
                graph.NNI_Link_ID = row[12]
                graph.VNNI_ID = row[13]
                graph.VLink_Circuit_ID = row[14]
                graph.STAG_Ranges = row[15]
                graph.max_bandwidth = int(row[16])  * 1000000
                RowsToUpdate.append(graph)
            else:
                print(f"POI:{row[5]} is missing")
        await Graphs.bulk_update(RowsToUpdate, fields=["AGVL", "Optus_STag", "NBN_STag","VTag_VLAN_ID", "CSA", "OPTUS_CVC","NBN_CVC", "OON", "ORD_VLINK", "ORD_CVC", "NNI_Link_ID", "VNNI_ID", "VLink_Circuit_ID", "STAG_Ranges" , "max_bandwidth"])
    return None

import yaml

with open("config.yaml", "r") as ymlfile:
    cfg = yaml.load(ymlfile, Loader=yaml.Loader)

run(start(cfg['databaseURL']))


# 0 : 'Project Reference - VNNIQLD', 
# 1 : 'AGVL', 
# 2 : 'Optus S-Tag', 
# 3 : 'NBN S-Tag',
# 4 : 'V-Tag VLAN ID',
# 5 : 'POI', 
# 6 : 'CSA', 
# 7 : 'Optus CVC', 
# 8 : 'NBN CVC', 
# 9 : 'OON', 
# 10: 'ORD VNNI', 
# 11: 'ORD CVC', 
# 12: 'NNI Link ID', 
# 13: 'V-NNI ID', 
# 14: 'V Link Circuit ID', 
# 15: 'S-TAG Ranges', 
# 16: 'Bandwidth (Mbps)'

